from __future__ import annotations

import argparse
import hashlib
import json
import re
from collections import Counter
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable

import openpyxl


DEFAULT_SOURCE = Path(
    "../唐长安城知识图谱数据工程/05-Excel数据/01-当前总表/"
    "BATCH-MASTER-0001_A_卷一至卷四与大事记29条_联合总表_20260823.xlsx"
)
DEFAULT_OUTPUT = Path("data/generated")
APPROVED_STATUSES = {"审核通过", "已审核", "已发布", "通过", "published"}
# 本项目已确认的 108 坊口径：排除重复/错误的崇仁坊，并保留两条有坐标的待复核坊。
EXCLUDED_ENTITY_KEYS = {"TC-FANG-0109"}
INCLUDED_DRAFT_ENTITY_KEYS = {"TC-FANG-0052", "TC-FANG-0055"}
BLOCKED_MARKERS = {"示例", "待补"}

SOURCE_REQUIRED = {"source_key", "title", "source_type"}
CHUNK_REQUIRED = {"chunk_key", "source_key", "content"}
ENTITY_REQUIRED = {"entity_key", "entity_type", "name"}
RELATION_REQUIRED = {"relation_key", "source_key", "target_key", "relation_type"}
RESEARCH_CLAIM_REQUIRED = {"claim_key", "claim_title", "claim_text"}

ENTITY_TYPES_WITH_LOCATION = {
    "AREA", "BUILDING", "CITY_GATE", "FANG", "GATE", "MARKET", "PALACE",
    "PLACE", "ROAD", "TERRAIN", "WATER", "区域", "坊", "地点", "地形",
    "城门", "宫殿", "建筑", "市场", "水体", "道路",
}


def clean_header(value: Any) -> str:
    return str(value or "").rstrip("*").strip()


def json_value(value: Any) -> Any:
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return value


def compact(obj: dict[str, Any]) -> dict[str, Any]:
    return {k: json_value(v) for k, v in obj.items() if v not in (None, "", [])}


def split_values(value: Any) -> list[str]:
    if value in (None, ""):
        return []
    return [part.strip() for part in re.split(r"[,;；，\n]+", str(value)) if part.strip()]


def format_source_display(source: dict[str, Any]) -> str:
    """Return the single bibliographic form that the game may display."""
    title = str(source.get("title") or "").strip()
    author = str(source.get("author") or "").strip()
    if not title:
        return ""
    formatted_title = title if title.startswith("《") else f"《{title}》"
    return f"{author}{formatted_title}" if author else formatted_title


def status(value: Any) -> str:
    return "published" if str(value or "").strip() in APPROVED_STATUSES else "draft"


def is_blocked(row: dict[str, Any]) -> bool:
    text = " ".join(str(v or "") for v in row.values())
    return any(marker in text for marker in BLOCKED_MARKERS)


def row_records(sheet: Any) -> tuple[int, list[str], list[dict[str, Any]]]:
    """Find the first header-like row, then return non-empty records below it."""
    best_row = 1
    best_headers: list[str] = []
    for row_number, cells in enumerate(sheet.iter_rows(min_row=1, max_row=12), start=1):
        headers = [clean_header(cell.value) for cell in cells]
        score = len([h for h in headers if h])
        known = len(
            set(headers)
            & (SOURCE_REQUIRED | CHUNK_REQUIRED | ENTITY_REQUIRED | RELATION_REQUIRED)
        )
        if (known, score) > (
            len(set(best_headers) & (SOURCE_REQUIRED | CHUNK_REQUIRED | ENTITY_REQUIRED | RELATION_REQUIRED)),
            len([h for h in best_headers if h]),
        ):
            best_row, best_headers = row_number, headers

    records: list[dict[str, Any]] = []
    for values in sheet.iter_rows(min_row=best_row + 1, values_only=True):
        record = {
            header: json_value(values[index] if index < len(values) else None)
            for index, header in enumerate(best_headers)
            if header
        }
        if any(value not in (None, "") for value in record.values()):
            records.append(record)
    return best_row, best_headers, records


def first_value(row: dict[str, Any], names: Iterable[str], default: Any = "") -> Any:
    for name in names:
        value = row.get(name)
        if value not in (None, ""):
            return value
    return default


def parse_number(value: Any) -> float | None:
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    if value in (None, ""):
        return None
    match = re.search(r"-?\d+(?:\.\d+)?", str(value))
    return float(match.group()) if match else None


def parse_pair(value: Any) -> tuple[float, float] | None:
    if value in (None, ""):
        return None
    nums = re.findall(r"-?\d+(?:\.\d+)?", str(value))
    if len(nums) >= 2:
        return float(nums[0]), float(nums[1])
    return None


def coordinate_from_row(row: dict[str, Any]) -> tuple[float, float] | None:
    pair = first_value(
        row,
        ("project_coordinate", "project_coordinates", "项目坐标", "项目坐标(x,y)", "坐标"),
        None,
    )
    parsed = parse_pair(pair)
    if parsed:
        return parsed
    x = parse_number(first_value(row, ("grid_x", "project_x", "项目x", "项目X", "x", "X"), None))
    y = parse_number(first_value(row, ("grid_y", "project_y", "项目y", "项目Y", "y", "Y"), None))
    return (x, y) if x is not None and y is not None else None


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def export(source: Path, output: Path, include_drafts: bool = False) -> dict[str, Any]:
    if not source.is_file():
        raise FileNotFoundError(f"找不到上游总表：{source}")

    workbook = openpyxl.load_workbook(source, read_only=True, data_only=True)
    inventory: list[dict[str, Any]] = []
    sources: dict[str, dict[str, Any]] = {}
    chunks: dict[str, dict[str, Any]] = {}
    entities: dict[str, dict[str, Any]] = {}
    relations_by_triple: dict[tuple[str, str, str], dict[str, Any]] = {}
    coordinate_overlays: dict[str, tuple[float, float]] = {}
    timeline_by_key: dict[str, dict[str, Any]] = {}
    research_claims_by_key: dict[str, dict[str, Any]] = {}
    skipped_blocked = 0
    skipped_drafts = 0

    parsed_sheets: list[tuple[str, list[str], list[dict[str, Any]]]] = []
    for sheet in workbook.worksheets:
        header_row, headers, rows = row_records(sheet)
        header_set = set(headers)
        inventory.append(
            {
                "sheet": sheet.title,
                "header_row": header_row,
                "row_count": len(rows),
                "headers": [h for h in headers if h],
            }
        )
        parsed_sheets.append((sheet.title, headers, rows))

        if RESEARCH_CLAIM_REQUIRED <= header_set:
            for row in rows:
                key = str(row.get("claim_key") or "").strip()
                if key:
                    research_claims_by_key[key] = {field: json_value(row.get(field)) for field in headers if field}

        for row in rows:
            entity_key = str(first_value(row, ("entity_key", "place_key", "fang_key", "building_key", "gate_key", "road_key", "canal_key", "terrain_key", "object_key"), "")).strip()
            coords = coordinate_from_row(row)
            if entity_key and coords:
                coordinate_overlays[entity_key] = coords

        if SOURCE_REQUIRED <= header_set:
            for row in rows:
                key = str(row.get("source_key") or "").strip()
                if not key:
                    continue
                sources[key] = compact({"source_key": key, **row, "status": status(row.get("review_status"))})

        if CHUNK_REQUIRED <= header_set:
            for row in rows:
                key = str(row.get("chunk_key") or "").strip()
                if not key:
                    continue
                chunks[key] = compact({"chunk_key": key, **row, "status": status(row.get("review_status"))})

        if ENTITY_REQUIRED <= header_set:
            for row in rows:
                key = str(row.get("entity_key") or "").strip()
                if not key:
                    continue
                if key in EXCLUDED_ENTITY_KEYS:
                    continue
                if is_blocked(row):
                    skipped_blocked += 1
                    continue
                item_status = status(row.get("review_status"))
                if item_status != "published" and key not in INCLUDED_DRAFT_ENTITY_KEYS and not include_drafts:
                    skipped_drafts += 1
                    continue
                item = compact(
                    {
                        "key": key,
                        "type": row.get("entity_type"),
                        "name": row.get("name"),
                        "trad": row.get("name_traditional"),
                        "aliases": split_values(row.get("aliases")),
                        "category": row.get("category"),
                        "period_start": row.get("period_start"),
                        "period_end": row.get("period_end"),
                        "description": row.get("description"),
                        "location": row.get("location_text"),
                        "source_key": row.get("source_key"),
                        "evidence_ids": split_values(row.get("chunk_key")),
                        "quote": row.get("description_quote"),
                        "review_status": item_status,
                        "confidence": row.get("confidence"),
                    }
                )
                if key in entities:
                    old = entities[key]
                    old.update({k: v for k, v in item.items() if v not in (None, "", [])})
                    old["evidence_ids"] = sorted(set(old.get("evidence_ids", []) + item.get("evidence_ids", [])))
                else:
                    entities[key] = item

        if RELATION_REQUIRED <= header_set:
            for row in rows:
                if is_blocked(row):
                    skipped_blocked += 1
                    continue
                item_status = status(row.get("review_status"))
                if item_status != "published" and not include_drafts:
                    skipped_drafts += 1
                    continue
                source_key = str(row.get("source_key") or "").strip()
                target_key = str(row.get("target_key") or "").strip()
                predicate = str(row.get("relation_type") or "").strip()
                if not source_key or not target_key or not predicate:
                    continue
                if source_key in EXCLUDED_ENTITY_KEYS or target_key in EXCLUDED_ENTITY_KEYS:
                    continue
                triple = (source_key, predicate, target_key)
                evidence = split_values(row.get("chunk_key"))
                if triple in relations_by_triple:
                    old = relations_by_triple[triple]
                    old["evidence_ids"] = sorted(set(old.get("evidence_ids", []) + evidence))
                else:
                    relations_by_triple[triple] = compact(
                        {
                            "key": row.get("relation_key"),
                            "source_id": source_key,
                            "predicate": predicate,
                            "label": row.get("relation_label"),
                            "target_id": target_key,
                            "period_start": row.get("period_start"),
                            "period_end": row.get("period_end"),
                            "evidence_ids": evidence,
                            "evidence_quote": row.get("evidence_quote"),
                            "evidence_type": row.get("evidence_type"),
                            "review_status": item_status,
                            "confidence": row.get("confidence"),
                            "note": row.get("note"),
                        }
                    )

    # Detect timeline-like business sheets independently of the universal entity table.
    for sheet_name, headers, rows in parsed_sheets:
        if not any(token in sheet_name for token in ("大事记", "时间", "历史事件", "timeline")):
            continue
        for index, row in enumerate(rows, start=1):
            year = parse_number(first_value(row, ("year", "year_ce", "event_year", "年份", "公元年份"), None))
            title = first_value(row, ("title", "event_name", "name", "事件名", "事件名称"), "")
            description = first_value(row, ("desc", "description", "summary", "historical_significance", "简述", "事件描述"), "")
            if year is None or not title:
                continue
            key = str(first_value(row, ("timeline_key", "event_key", "entity_key", "key", "事件编号"), f"EVENT-{int(year):04d}-{index:03d}"))
            source_keys = split_values(first_value(row, ("source_key", "source_ids", "来源编号"), ""))
            evidence_ids = split_values(first_value(row, ("chunk_key", "evidence_ids", "证据片段编号"), ""))
            related_ids = split_values(first_value(row, ("related_entity_keys", "related_point_ids", "关联实体编号"), ""))
            event_entity = compact(
                {
                    "key": key,
                    "type": "HistoricalEvent",
                    "name": title,
                    "description": description,
                    "period_start": int(year) if year.is_integer() else year,
                    "period_end": int(year) if year.is_integer() else year,
                    "source_key": source_keys[0] if len(source_keys) == 1 else source_keys,
                    "evidence_ids": evidence_ids,
                    "review_status": status(row.get("review_status")),
                    "confidence": row.get("confidence"),
                    "properties": compact({
                        field: value
                        for field, value in row.items()
                        if field not in {"timeline_key", "source_key", "chunk_key", "review_status", "confidence"}
                    }),
                }
            )
            entities[key] = event_entity
            timeline_by_key[key] = {
                **event_entity,
                "id": key,
                "year": int(year) if year.is_integer() else year,
                "title": title,
                "desc": description,
                "source_ids": source_keys,
                "related_point_ids": related_ids,
            }

    # Preserve fields that exist only in specialty tables.  These tables use
    # type-specific key columns rather than the universal `entity_key` column.
    specialty_key_fields = (
        "fang_key", "road_key", "canal_key", "gate_key", "terrain_key", "building_key",
    )
    shared_entity_fields = ENTITY_REQUIRED | {
        "name_traditional", "aliases", "category", "period_start", "period_end",
        "description", "location_text", "source_key", "chunk_key", "confidence",
        "review_status", "description_quote", "x", "y", "coordinate_system",
    }
    for sheet_name, headers, rows in parsed_sheets:
        key_field = next((field for field in specialty_key_fields if field in headers), None)
        if key_field is None:
            continue
        for row in rows:
            key = str(row.get(key_field) or "").strip()
            if key not in entities:
                continue
            properties = compact({
                field: value
                for field, value in row.items()
                if field not in shared_entity_fields and field != key_field
            })
            if properties:
                entities[key].setdefault("properties", {}).update(properties)
    for key, coords in coordinate_overlays.items():
        if key in entities:
            entities[key]["grid_x"], entities[key]["grid_y"] = coords

    entity_ids = set(entities)
    chunk_ids = set(chunks)
    source_ids = set(sources)
    relations = list(relations_by_triple.values())
    orphan_relations = [
        item.get("key", "")
        for item in relations
        if item["source_id"] not in entity_ids or item["target_id"] not in entity_ids
    ]
    bad_relation_evidence = [
        item.get("key", "")
        for item in relations
        if any(evidence not in chunk_ids for evidence in item.get("evidence_ids", []))
    ]
    bad_chunk_sources = [
        item.get("chunk_key", "")
        for item in chunks.values()
        if item.get("source_key") and item["source_key"] not in source_ids
    ]

    valid_relations = [
        item for item in relations
        if item["source_id"] in entity_ids and item["target_id"] in entity_ids
    ]

    points: list[dict[str, Any]] = []
    for item in entities.values():
        entity_type = str(item.get("type", ""))
        if "grid_x" not in item or "grid_y" not in item:
            continue
        if entity_type not in ENTITY_TYPES_WITH_LOCATION and entity_type.upper() not in ENTITY_TYPES_WITH_LOCATION and not any(
            token in entity_type for token in ("地点", "区域", "坊", "宫", "门", "路", "渠", "水", "建筑", "地形")
        ):
            continue
        point = dict(item)
        point["zone"] = item.get("category", "")
        point["source_ids"] = split_values(item.get("source_key"))
        # `source` is deliberately generated from the source table, rather than
        # from book titles that happen to occur inside a quoted passage.  The
        # Godot prompt uses this field as its only displayable citation.
        point["source"] = "；".join(
            display
            for source_id in point["source_ids"]
            if (display := format_source_display(sources.get(source_id, {})))
        )
        points.append(compact(point))

    timeline = sorted(timeline_by_key.values(), key=lambda item: (item["year"], item["id"]))
    research_claims = sorted(research_claims_by_key.values(), key=lambda item: item["claim_key"])
    type_counts = Counter(str(item.get("type", "未分类")) for item in entities.values())
    validation = {
        "source_count": len(sources),
        "source_chunk_count": len(chunks),
        "entity_count": len(entities),
        "relation_count": len(valid_relations),
        "point_count": len(points),
        "timeline_count": len(timeline),
        "research_claim_count": len(research_claims),
        "entity_type_counts": dict(sorted(type_counts.items())),
        "orphan_relations": orphan_relations,
        "bad_relation_evidence": bad_relation_evidence,
        "bad_chunk_sources": bad_chunk_sources,
        "skipped_blocked_rows": skipped_blocked,
        "skipped_draft_rows": skipped_drafts,
        "include_drafts": include_drafts,
        "ok": not (orphan_relations or bad_relation_evidence or bad_chunk_sources),
    }
    manifest = {
        "source_file": str(source.resolve()),
        "source_filename": source.name,
        "source_sha256": sha256(source),
        "exported_at": datetime.now().astimezone().isoformat(timespec="seconds"),
        "workbook_sheets": inventory,
        "outputs": [
            "sources.json", "source_chunks.json", "entities.json", "relations.json",
            "changan_points.json", "history_timeline.json", "research_claims.json",
            "validation_report.json",
        ],
    }

    write_json(output / "sources.json", {"sources": list(sources.values())})
    write_json(output / "source_chunks.json", {"source_chunks": list(chunks.values())})
    write_json(output / "entities.json", {"entities": list(entities.values())})
    write_json(output / "relations.json", {"relations": valid_relations})
    write_json(output / "changan_points.json", {"points": points})
    write_json(output / "history_timeline.json", {"timeline": timeline})
    write_json(output / "research_claims.json", {"research_claims": research_claims})
    write_json(output / "validation_report.json", validation)
    write_json(output / "import_manifest.json", manifest)
    return validation


def main() -> int:
    parser = argparse.ArgumentParser(description="将唐长安城知识图谱联合总表导出为 Godot JSON。")
    parser.add_argument("--source", type=Path, default=DEFAULT_SOURCE, help="联合总表 .xlsx 路径")
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT, help="生成目录")
    parser.add_argument("--include-drafts", action="store_true", help="同时导出未审核数据（默认关闭）")
    args = parser.parse_args()
    report = export(args.source.resolve(), args.output.resolve(), args.include_drafts)
    print(json.dumps(report, ensure_ascii=False, indent=2))
    return 0 if report["ok"] else 2


if __name__ == "__main__":
    raise SystemExit(main())
